from __future__ import annotations

import json
import re
from pathlib import Path

import fitz

from vz_search.domain.entities import ExtractedPerson
from vz_search.infrastructure.ai.prompts import EXTRACTION_PROMPT


def parse_persons_json(raw: str) -> list[ExtractedPerson]:
    text = raw.strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\s*", "", text)
        text = re.sub(r"\s*```$", "", text)

    payload = json.loads(text)
    items = payload.get("personas") or payload.get("persons") or payload
    if not isinstance(items, list):
        return []

    persons: list[ExtractedPerson] = []
    for item in items:
        if not isinstance(item, dict):
            continue
        full_name = str(item.get("full_name") or item.get("nombre_completo") or "").strip()
        if not full_name:
            first = str(item.get("nombre") or item.get("first_name") or "").strip()
            last = str(item.get("apellido") or item.get("last_name") or "").strip()
            full_name = f"{first} {last}".strip()
        if not full_name:
            continue

        persons.append(
            ExtractedPerson(
                full_name=full_name,
                hospital=_null_str(item.get("hospital")),
                state=_null_str(item.get("state") or item.get("estado")),
                cedula=_null_str(item.get("cedula") or item.get("cédula")),
                age=_null_str(item.get("age") or item.get("edad")),
                condition=_null_str(item.get("condition") or item.get("condicion")),
                notes=_null_str(item.get("notes") or item.get("notas")),
            )
        )
    return persons


def _null_str(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.lower() in {"null", "none", "n/a", "na", "-"}:
        return None
    return text


def pdf_pages_as_png_bytes(path: Path, max_pages: int = 15) -> list[bytes]:
    doc = fitz.open(path)
    images: list[bytes] = []
    for index, page in enumerate(doc):
        if index >= max_pages:
            break
        pix = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
        images.append(pix.tobytes("png"))
    doc.close()
    return images


def load_image_bytes(path: Path) -> list[bytes]:
    return [path.read_bytes()]


def read_text_file(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="ignore")


def read_docx_text(path: Path) -> str:
    import re
    import zipfile

    with zipfile.ZipFile(path) as archive:
        xml_content = archive.read("word/document.xml").decode("utf-8", errors="ignore")
    text = re.sub(r"<w:tab/>", "\t", xml_content)
    text = re.sub(r"</w:p>", "\n", text)
    text = re.sub(r"<[^>]+>", " ", text)
    return re.sub(r"[ \t]+", " ", text).strip()


def read_xlsx_text(path: Path) -> str:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    rows: list[str] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(values_only=True):
            cells = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
            if cells:
                rows.append(" | ".join(cells))
    workbook.close()
    return "\n".join(rows)
